"""
Baseline Engine for Direct LLM Prompting

This module implements a baseline approach that directly prompts the LLM
with database content loaded from xlsx files to answer natural language questions.
"""

import os
import time
import logging
from typing import Optional, List, Dict, Any, Union, Tuple
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from semantic_parser.llm_client import AzureOpenAIClient, Message


EXCEL_FILE_DIR = "/home/jiuding/computational_thinking/semantic_parser/src/baseline/db.xlsx"


class SmartExcelLoader:
    """
    Smart Excel loader that can detect multiple tables within a sheet.
    
    Handles:
    - Multiple tables per sheet (spread across different areas)
    - Empty rows between tables
    - Colored header rows for table detection
    """
    
    def __init__(
        self,
        min_table_cols: int = 2,
        min_table_rows: int = 2,
        header_color_threshold: float = 0.3,
        verbose: bool = False,
    ):
        """
        Initialize the smart Excel loader.
        
        Args:
            min_table_cols: Minimum columns for a valid table
            min_table_rows: Minimum data rows for a valid table (excluding header)
            header_color_threshold: Fraction of cells that must be colored to consider a header row
            verbose: Whether to print debug information
        """
        self.min_table_cols = min_table_cols
        self.min_table_rows = min_table_rows
        self.header_color_threshold = header_color_threshold
        self.verbose = verbose
    
    def _is_cell_colored(self, cell) -> bool:
        """Check if a cell has a background color (not white/no fill)."""
        if cell.fill is None:
            return False
        
        fill = cell.fill
        if fill.fill_type is None or fill.fill_type == 'none':
            return False
        
        # Check foreground color
        if fill.fgColor:
            # Check if it's not white (FFFFFFFF) or no color (00000000)
            color = fill.fgColor
            if color.type == 'rgb' and color.rgb:
                rgb = color.rgb
                # Skip white and transparent
                if rgb in ('FFFFFFFF', '00000000', 'FFFFFF', '000000'):
                    return False
                return True
            elif color.type == 'theme':
                # Theme colors are typically non-white
                return True
            elif color.type == 'indexed':
                # Indexed color 64 is often "no color"
                return color.indexed != 64
        
        return False
    
    def _is_header_row(self, ws, row_idx: int, start_col: int, end_col: int) -> bool:
        """Check if a row appears to be a header row based on cell coloring."""
        colored_count = 0
        non_empty_count = 0
        
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                non_empty_count += 1
                if self._is_cell_colored(cell):
                    colored_count += 1
        
        if non_empty_count == 0:
            return False
        
        # Consider it a header if enough cells are colored
        color_ratio = colored_count / non_empty_count
        return color_ratio >= self.header_color_threshold
    
    def _is_row_empty(self, ws, row_idx: int, start_col: int, end_col: int) -> bool:
        """Check if a row is empty (all cells are None or whitespace)."""
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                return False
        return True
    
    def _find_table_boundaries(self, ws, start_row: int, start_col: int) -> Optional[Tuple[int, int, int, int]]:
        """
        Find the boundaries of a table starting from the given position.
        
        Returns:
            Tuple of (start_row, end_row, start_col, end_col) or None if no valid table
        """
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Find the extent of columns (look for continuous non-empty header cells)
        end_col = start_col
        for col_idx in range(start_col, max_col + 1):
            cell = ws.cell(row=start_row, column=col_idx)
            if cell.value is None or not str(cell.value).strip():
                # Allow one empty column, but stop at two consecutive empties
                next_cell = ws.cell(row=start_row, column=col_idx + 1) if col_idx < max_col else None
                if next_cell is None or next_cell.value is None or not str(next_cell.value).strip():
                    break
            end_col = col_idx
        
        if end_col - start_col + 1 < self.min_table_cols:
            return None
        
        # Find the extent of rows (until we hit multiple empty rows)
        end_row = start_row
        consecutive_empty = 0
        
        for row_idx in range(start_row + 1, max_row + 1):
            if self._is_row_empty(ws, row_idx, start_col, end_col):
                consecutive_empty += 1
                if consecutive_empty >= 2:
                    break
            else:
                consecutive_empty = 0
                end_row = row_idx
        
        # Check minimum rows requirement
        if end_row - start_row < self.min_table_rows:
            return None
        
        return (start_row, end_row, start_col, end_col)
    
    def _extract_table(self, ws, bounds: Tuple[int, int, int, int], table_name: str) -> pd.DataFrame:
        """Extract a DataFrame from the given boundaries."""
        start_row, end_row, start_col, end_col = bounds
        
        # Extract header row
        headers = []
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col_idx)
            header = str(cell.value) if cell.value is not None else f"Column_{col_idx}"
            headers.append(header.strip())
        
        # Make headers unique
        seen = {}
        unique_headers = []
        for h in headers:
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                unique_headers.append(h)
        
        # Extract data rows
        data = []
        for row_idx in range(start_row + 1, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            data.append(row_data)
        
        df = pd.DataFrame(data, columns=unique_headers)
        
        # Drop rows that are completely empty
        df = df.dropna(how='all')
        
        return df
    
    def _detect_tables_in_sheet(self, ws) -> List[Tuple[str, Tuple[int, int, int, int]]]:
        """
        Detect all tables in a worksheet.
        
        Returns:
            List of (table_name, boundaries) tuples
        """
        tables = []
        processed_cells = set()
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Scan for header rows (colored rows with content)
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                # Skip if already part of a table
                if (row_idx, col_idx) in processed_cells:
                    continue
                
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Check if this could be a table header start
                if cell.value is not None and str(cell.value).strip():
                    # Check if this row looks like a header
                    # First, find potential column range
                    temp_end_col = col_idx
                    for c in range(col_idx, max_col + 1):
                        test_cell = ws.cell(row=row_idx, column=c)
                        if test_cell.value is None or not str(test_cell.value).strip():
                            next_cell = ws.cell(row=row_idx, column=c + 1) if c < max_col else None
                            if next_cell is None or next_cell.value is None:
                                break
                        temp_end_col = c
                    
                    # Check if it's a header row (colored or has data below)
                    is_header = self._is_header_row(ws, row_idx, col_idx, temp_end_col)
                    
                    # Even if not colored, check if there's data below
                    if not is_header:
                        has_data_below = False
                        for test_row in range(row_idx + 1, min(row_idx + 3, max_row + 1)):
                            if not self._is_row_empty(ws, test_row, col_idx, temp_end_col):
                                has_data_below = True
                                break
                        is_header = has_data_below
                    
                    if is_header:
                        bounds = self._find_table_boundaries(ws, row_idx, col_idx)
                        if bounds:
                            start_r, end_r, start_c, end_c = bounds
                            
                            # Mark cells as processed
                            for r in range(start_r, end_r + 1):
                                for c in range(start_c, end_c + 1):
                                    processed_cells.add((r, c))
                            
                            # Generate table name from first cell or row index
                            first_cell_value = str(cell.value).strip()[:30]
                            table_name = f"Table_{row_idx}_{first_cell_value}"
                            table_name = table_name.replace(" ", "_").replace("-", "_")
                            
                            tables.append((table_name, bounds))
        
        return tables
    
    def load_workbook(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Load all tables from an Excel workbook.
        
        Args:
            file_path: Path to the xlsx file
            
        Returns:
            Dictionary mapping table names to DataFrames
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        wb = load_workbook(file_path, data_only=True)
        tables = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            if self.verbose:
                logging.info(f"Processing sheet: {sheet_name}")
            
            # Detect tables in this sheet
            detected_tables = self._detect_tables_in_sheet(ws)
            
            if self.verbose:
                logging.info(f"  Found {len(detected_tables)} tables")
            
            for table_name, bounds in detected_tables:
                # Create unique table name including sheet name
                full_table_name = f"{sheet_name}_{table_name}"
                full_table_name = full_table_name.replace(" ", "_").replace("-", "_")
                
                # Extract the table
                df = self._extract_table(ws, bounds, full_table_name)
                
                if len(df) >= self.min_table_rows:
                    tables[full_table_name] = df
                    
                    if self.verbose:
                        logging.info(f"    Extracted '{full_table_name}': {len(df)} rows, {len(df.columns)} cols")
        
        wb.close()
        return tables


@dataclass
class BaselineResult:
    """Result from baseline engine query."""
    query: str
    answer: str
    tables_used: List[str]
    total_rows: int
    duration_seconds: float
    timestamp: datetime = field(default_factory=datetime.now)
    usage: Optional[Dict[str, int]] = None
    
    def __str__(self) -> str:
        return f"Query: {self.query}\nAnswer: {self.answer}"


class BaselineEngine:
    """
    Baseline engine that directly prompts LLM with xlsx data.
    
    This serves as a comparison baseline for more sophisticated
    approaches like ReACT-based reasoning.
    """
    
    def __init__(
        self,
        llm_client: AzureOpenAIClient,
        xlsx_paths: Optional[Union[str, List[str]]] = [EXCEL_FILE_DIR],
        max_rows_per_table: int = 1000,
        smart_load: bool = True,
        verbose: bool = False,
    ):
        """
        Initialize the baseline engine.
        
        Args:
            llm_client: Azure OpenAI client for LLM calls
            xlsx_paths: Path(s) to xlsx file(s). Can be a single path,
                       a list of paths, or a directory containing xlsx files.
            max_rows_per_table: Maximum rows to include per table (to avoid token limits)
            smart_load: If True, uses SmartExcelLoader to detect multiple tables
                       per sheet and handle colored headers. If False, loads
                       each sheet as a single table.
            verbose: Whether to print debug information
        """
        self.llm_client = llm_client
        self.max_rows_per_table = max_rows_per_table
        self.smart_load = smart_load
        self.verbose = verbose
        
        # Store loaded dataframes
        self.tables: Dict[str, pd.DataFrame] = {}
        
        # Initialize smart loader if needed
        if smart_load:
            self.smart_loader = SmartExcelLoader(verbose=verbose)
        
        # Load xlsx files if provided
        if xlsx_paths:
            self.load_xlsx(xlsx_paths)
    
    def load_xlsx(self, paths: Union[str, List[str]]) -> None:
        """
        Load xlsx file(s) into memory.
        
        Args:
            paths: Path(s) to xlsx file(s). Can be:
                   - A single xlsx file path
                   - A list of xlsx file paths
                   - A directory containing xlsx files
        """
        if isinstance(paths, str):
            path = Path(paths)
            if path.is_dir():
                # Load all xlsx files in directory
                xlsx_files = list(path.glob("*.xlsx"))
                if not xlsx_files:
                    raise ValueError(f"No xlsx files found in directory: {paths}")
                paths = [str(f) for f in xlsx_files]
            else:
                paths = [paths]
        
        for file_path in paths:
            self._load_single_xlsx(file_path)
        
        if self.verbose:
            logging.info(f"Loaded {len(self.tables)} tables: {list(self.tables.keys())}")
    
    def _load_single_xlsx(self, file_path: str) -> None:
        """Load a single xlsx file, potentially with multiple sheets and tables."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if not path.suffix.lower() == '.xlsx':
            raise ValueError(f"File must be .xlsx format: {file_path}")
        
        base_name = path.stem
        
        if self.smart_load:
            # Use smart loader to detect multiple tables per sheet
            loaded_tables = self.smart_loader.load_workbook(file_path)
            
            for table_name, df in loaded_tables.items():
                # Add file prefix to table name
                full_name = f"{base_name}_{table_name}"
                full_name = full_name.replace(" ", "_").replace("-", "_")
                
                # Convert column names to strings
                df.columns = [str(col) for col in df.columns]
                
                self.tables[full_name] = df
                
                if self.verbose:
                    print(f"Loaded table '{full_name}': {len(df)} rows, {len(df.columns)} columns")
        else:
            # Simple loading: one table per sheet
            xlsx = pd.ExcelFile(file_path)
            
            for sheet_name in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet_name)
                
                # Convert column names to strings (handles datetime columns from Excel)
                df.columns = [str(col) for col in df.columns]
                
                # Generate table name from file and sheet
                if len(xlsx.sheet_names) > 1:
                    table_name = f"{base_name}_{sheet_name}"
                else:
                    table_name = base_name
                
                # Clean up table name
                table_name = table_name.replace(" ", "_").replace("-", "_")
                
                self.tables[table_name] = df
                
                if self.verbose:
                    logging.info(f"Loaded table '{table_name}': {len(df)} rows, {len(df.columns)} columns")
    
    def add_dataframe(self, name: str, df: pd.DataFrame) -> None:
        """
        Add a pandas DataFrame directly as a table.
        
        Args:
            name: Name for the table
            df: DataFrame to add
        """
        self.tables[name] = df
        if self.verbose:
            logging.info(f"Added table '{name}': {len(df)} rows, {len(df.columns)} columns")
    
    def _format_tables_for_prompt(self) -> str:
        """Format all loaded tables as a string for the LLM prompt."""
        table_strings = []
        
        for table_name, df in self.tables.items():
            # Limit rows if needed
            if len(df) > self.max_rows_per_table:
                df_display = df.head(self.max_rows_per_table)
                truncated_note = f"\n(Showing first {self.max_rows_per_table} of {len(df)} rows)"
            else:
                df_display = df
                truncated_note = ""
            
            # Format as markdown table
            table_md = df_display.to_markdown(index=False)
            
            # Convert column names to strings (handles datetime columns)
            column_names = [str(col) for col in df.columns.tolist()]
            
            # Add table info
            table_info = f"""
## Table: {table_name}
Columns: {', '.join(column_names)}
Total Rows: {len(df)}{truncated_note}

{table_md}
"""
            table_strings.append(table_info)
        
        return "\n\n---\n".join(table_strings)
    
    def _build_prompt(self, query: str, context: Optional[str] = None) -> str:
        """Build the prompt for the LLM."""
        tables_content = self._format_tables_for_prompt()
        
        context_section = ""
        if context:
            context_section = f"""
## Additional Context
{context}
"""
        
        prompt = f"""You are a data analyst assistant. Answer the following question based on the database tables provided below.

## Question
{query}
{context_section}
## Database Tables
{tables_content}

## Instructions
1. Analyze the provided tables to answer the question
2. If calculations are needed, show your work
3. Be precise with numbers and use the exact values from the data
4. If the question cannot be answered with the available data, explain why

## Answer
"""
        return prompt
    
    def query(
        self,
        query: str,
        context: Optional[str] = None,
        system_prompt: Optional[str] = None,
    ) -> BaselineResult:
        """
        Answer a question using direct LLM prompting with the loaded data.
        
        Args:
            query: Natural language question to answer
            context: Optional additional context for the query
            system_prompt: Optional custom system prompt
            
        Returns:
            BaselineResult containing the answer and metadata
        """
        if not self.tables:
            raise ValueError("No tables loaded. Call load_xlsx() first.")
        
        start_time = time.time()
        
        # Build the prompt
        user_prompt = self._build_prompt(query, context)
        
        # Default system prompt
        if system_prompt is None:
            system_prompt = """You are an expert data analyst. Your task is to analyze database tables and answer questions accurately.
Always base your answers on the actual data provided. Be precise with numbers and calculations.
If you need to perform calculations, show your reasoning step by step."""
        
        # Create messages
        messages = [
            Message(role="system", content=system_prompt),
            Message(role="user", content=user_prompt)
        ]
        
        if self.verbose:
            logging.info(f"Querying LLM with {len(user_prompt)} character prompt")
            print(f"\n{'='*60}")
            print(f"BASELINE QUERY: {query}")
            print(f"{'='*60}")
        
        # Call LLM
        response = self.llm_client.chat_completion(messages)
        
        duration = time.time() - start_time
        
        # Calculate total rows
        total_rows = sum(len(df) for df in self.tables.values())
        
        result = BaselineResult(
            query=query,
            answer=response.content.strip(),
            tables_used=list(self.tables.keys()),
            total_rows=total_rows,
            duration_seconds=duration,
            usage=response.usage,
        )
        
        if self.verbose:
            print(f"\nANSWER:\n{result.answer}")
            print(f"\nDuration: {duration:.2f}s")
            print(f"Tables used: {result.tables_used}")
            print(f"{'='*60}\n")
        
        return result
    
    def get_table_info(self) -> Dict[str, Dict[str, Any]]:
        """Get information about all loaded tables."""
        info = {}
        for name, df in self.tables.items():
            info[name] = {
                "columns": [str(col) for col in df.columns.tolist()],
                "dtypes": {str(k): str(v) for k, v in df.dtypes.to_dict().items()},
                "row_count": len(df),
                "sample": df.head(3).to_dict(orient="records"),
            }
        return info
    
    def clear_tables(self) -> None:
        """Clear all loaded tables."""
        self.tables = {}
        if self.verbose:
            logging.info("Cleared all tables")

